import mysql.connector
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

# 添加應用程式路徑
sys.path.insert(0, '/app')

# 導入資料庫配置
from backend.database.config import Config

# ==================== 資料庫連線設定 ====================
config = {
    'host': Config.LAB_DB_HOST,
    'port': Config.LAB_DB_PORT,
    'user': Config.LAB_DB_USERNAME,
    'password': Config.LAB_DB_PASSWORD,
    'database': Config.LAB_DB_NAME,
    'charset': 'utf8mb4'
}

# ==================== 連接資料庫並讀取資料 ====================
print("正在連接 LAB 資料庫...")
print(f"連線資訊: {Config.LAB_DB_HOST}:{Config.LAB_DB_PORT}/{Config.LAB_DB_NAME}")

try:
    conn = mysql.connector.connect(**config)
    print("✅ 資料庫連線成功！")
    
    # 讀取各統計表
    tables = {
        'login_stats': {
            'query': 'SELECT * FROM login_stats ORDER BY month, student_level',
            'instances_col': 'login_instances',
            'users_col': 'login_users',
            'cumulative_col': 'cumulative_users',
            'instances_label': '登入人次',
            'users_label': '登入人數'
        },
        'usage_stats': {
            'query': 'SELECT * FROM usage_stats ORDER BY month, student_level',
            'instances_col': 'usage_instances',
            'users_col': 'usage_users',
            'cumulative_col': 'cumulative_users',
            'instances_label': '使用人次',
            'users_label': '使用人數'
        },
        'coolebot_stats': {
            'query': 'SELECT * FROM coolebot_stats ORDER BY month, student_level',
            'instances_col': 'usage_instances',
            'users_col': 'usage_users',
            'cumulative_col': 'cumulative_users',
            'instances_label': '使用人次',
            'users_label': '使用人數'
        }
    }
    
    raw_dataframes = {}
    
    for table_name, table_info in tables.items():
        print(f"\n正在讀取 {table_name}...")
        df = pd.read_sql(table_info['query'], conn)
        raw_dataframes[table_name] = {'df': df, 'info': table_info}
        print(f"✅ {table_name} 讀取完成，共 {len(df)} 筆資料")
    
    # 讀取 ai_course_stats（格式不同）
    print(f"\n正在讀取 ai_course_stats...")
    ai_course_df = pd.read_sql('SELECT * FROM ai_course_stats ORDER BY month, ai_course_name, student_level', conn)
    raw_dataframes['ai_course_stats'] = {'df': ai_course_df, 'info': None}
    print(f"✅ ai_course_stats 讀取完成，共 {len(ai_course_df)} 筆資料")
    
    conn.close()
    print("\n資料庫連線已關閉")
    
except mysql.connector.Error as err:
    print(f"❌ 資料庫連線錯誤: {err}")
    sys.exit(1)
    
except Exception as e:
    print(f"❌ 其他錯誤: {e}")
    import traceback
    traceback.print_exc()
    sys.exit(1)

# ==================== 處理資料並建立透視表 ====================
def create_pivot_table(df, table_info):
    """將資料轉換為透視表格式"""
    levels = ['國小', '國中', '高中職', '其他']
    
    # 建立透視表
    pivot_instances = df.pivot_table(
        index='month', 
        columns='student_level', 
        values=table_info['instances_col'], 
        fill_value=0, 
        aggfunc='sum'
    )
    
    pivot_users = df.pivot_table(
        index='month', 
        columns='student_level', 
        values=table_info['users_col'], 
        fill_value=0, 
        aggfunc='sum'
    )
    
    # 確保所有學制都存在
    for level in levels:
        if level not in pivot_instances.columns:
            pivot_instances[level] = 0
        if level not in pivot_users.columns:
            pivot_users[level] = 0
    
    # 重新排序欄位
    pivot_instances = pivot_instances[levels].sort_index()
    pivot_users = pivot_users[levels].sort_index()
    
    # 計算總計
    pivot_instances['總計'] = pivot_instances.sum(axis=1)
    pivot_users['總計'] = pivot_users[levels].sum(axis=1)
    
    # 處理累計人數（只保留總計的累計值）
    if table_info['cumulative_col'] in df.columns:
        cumulative_data = df[df['student_level'] == '總計'].set_index('month')[table_info['cumulative_col']]
        pivot_users['累計'] = pivot_users.index.map(lambda m: int(cumulative_data.get(m, 0)))
    else:
        pivot_users['累計'] = 0
    
    return pivot_instances, pivot_users, table_info

# ==================== 建立 Excel 檔案 ====================
output_file = f'統計資料匯總_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
print(f"\n正在建立 Excel 檔案：{output_file}")

wb = Workbook()
# 移除預設的工作表
wb.remove(wb.active)

# 顏色定義
header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # 淺藍色
month_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 淺黃色
total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # 淺藍色
cumulative_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 淺綠色
header_font = Font(name='微軟正黑體', size=11, bold=True)
data_font = Font(name='微軟正黑體', size=10)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# 處理前三個統計表（格式相同）
for table_name, table_data in raw_dataframes.items():
    if table_name == 'ai_course_stats':
        continue
    
    df = table_data['df']
    table_info = table_data['info']
    
    if df.empty:
        print(f"⚠️ {table_name} 沒有資料，跳過")
        continue
    
    print(f"\n正在處理 {table_name}...")
    
    # 建立透視表
    pivot_instances, pivot_users, info = create_pivot_table(df, table_info)
    
    # 建立工作表
    ws = wb.create_sheet(title=table_name)
    
    # 寫入標題列（第一行：主標題）
    row_idx = 1
    col_idx = 1
    
    # 月份欄位（合併 2 行）
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    cell = ws.cell(row=1, column=1)
    cell.value = '月份'
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border
    
    # 登入人次/使用人次（合併多欄）
    instances_start_col = 2
    instances_end_col = 2 + len(['國小', '國中', '高中職', '其他', '總計']) - 1
    ws.merge_cells(start_row=1, start_column=instances_start_col, end_row=1, end_column=instances_end_col)
    cell = ws.cell(row=1, column=instances_start_col)
    cell.value = info['instances_label']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border
    
    # 登入人數/使用人數（合併多欄）
    users_start_col = instances_end_col + 1
    users_end_col = users_start_col + len(['國小', '國中', '高中職', '其他', '總計', '累計']) - 1
    ws.merge_cells(start_row=1, start_column=users_start_col, end_row=1, end_column=users_end_col)
    cell = ws.cell(row=1, column=users_start_col)
    cell.value = info['users_label']
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border
    
    # 第二行：子標題
    row_idx = 2
    col_idx = 2
    
    # 人次子標題
    for level in ['國小', '國中', '高中職', '其他', '總計']:
        cell = ws.cell(row=2, column=col_idx)
        cell.value = f"{level}{info['instances_label']}"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        if level == '總計':
            cell.fill = total_fill
        col_idx += 1
    
    # 人數子標題
    for level in ['國小', '國中', '高中職', '其他', '總計', '累計']:
        cell = ws.cell(row=2, column=col_idx)
        if level == '累計':
            cell.value = f"累計{info['users_label']}"
        else:
            cell.value = f"{level}{info['users_label']}"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        if level == '總計':
            cell.fill = total_fill
        elif level == '累計':
            cell.fill = cumulative_fill
        col_idx += 1
    
    # 寫入資料
    row_idx = 3
    for month in sorted(pivot_instances.index):
        col_idx = 1
        
        # 月份（格式：YYYY/M）
        month_str = month.replace('-', '/')
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = month_str
        cell.fill = month_fill
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = thin_border
        col_idx += 1
        
        # 人次資料
        for level in ['國小', '國中', '高中職', '其他', '總計']:
            cell = ws.cell(row=row_idx, column=col_idx)
            value = int(pivot_instances.loc[month, level])
            cell.value = value if value > 0 else 0
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            if level == '總計':
                cell.fill = total_fill
            col_idx += 1
        
        # 人數資料
        for level in ['國小', '國中', '高中職', '其他', '總計', '累計']:
            cell = ws.cell(row=row_idx, column=col_idx)
            if level == '累計':
                value = int(pivot_users.loc[month, '累計']) if '累計' in pivot_users.columns else 0
                cell.fill = cumulative_fill
            else:
                value = int(pivot_users.loc[month, level])
                if level == '總計':
                    cell.fill = total_fill
            cell.value = value if value > 0 else 0
            cell.font = data_font
            cell.alignment = center_alignment
            cell.border = thin_border
            col_idx += 1
        
        row_idx += 1
    
    # 調整欄寬
    ws.column_dimensions['A'].width = 12  # 月份欄位
    for col_idx in range(2, col_idx):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 15
    
    # 凍結窗格
    ws.freeze_panes = 'A3'
    
    print(f"✅ {table_name} 工作表建立完成")

# 處理 ai_course_stats（不同格式）
if 'ai_course_stats' in raw_dataframes and not raw_dataframes['ai_course_stats']['df'].empty:
    print(f"\n正在處理 ai_course_stats...")
    df = raw_dataframes['ai_course_stats']['df']
    
    # 建立透視表：課程 x (月份, 學制)
    pivot_df = df.pivot_table(
        index='ai_course_name',
        columns=['month', 'student_level'],
        values='user_count',
        fill_value=0,
        aggfunc='sum'
    )
    
    ws = wb.create_sheet(title='ai_course_stats')
    
    # 寫入標題（簡化版）
    ws.cell(row=1, column=1, value='課程名稱').fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = center_alignment
    
    # 寫入月份和學制標題
    col_idx = 2
    for month, level in sorted(pivot_df.columns):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = f"{month}\n{level}"
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        col_idx += 1
    
    # 寫入資料
    row_idx = 2
    for course in pivot_df.index:
        ws.cell(row=row_idx, column=1, value=course).font = data_font
        col_idx = 2
        for month, level in sorted(pivot_df.columns):
            value = int(pivot_df.loc[course, (month, level)])
            ws.cell(row=row_idx, column=col_idx, value=value if value > 0 else '').font = data_font
            ws.cell(row=row_idx, column=col_idx).alignment = center_alignment
            col_idx += 1
        row_idx += 1
    
    # 調整欄寬
    ws.column_dimensions['A'].width = 35
    for col_idx in range(2, col_idx):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 10
    
    ws.freeze_panes = 'B2'
    print(f"✅ ai_course_stats 工作表建立完成")

# 儲存檔案
wb.save(output_file)
print(f"\n✅ Excel 檔案建立完成：{output_file}")
