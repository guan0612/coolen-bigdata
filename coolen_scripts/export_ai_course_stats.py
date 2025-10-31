import mysql.connector
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
import sys
import os

# 添加應用程式路徑
sys.path.insert(0, '/app')

# 導入資料庫配置
from backend.database.config import Config
from backend.database.service_lab import ServiceLAB
from backend.database.service_ncu import ServiceNCU

# ==================== 資料庫連線設定 ====================
config = {
    'host': Config.LAB_DB_HOST,
    'port': Config.LAB_DB_PORT,
    'user': Config.LAB_DB_USERNAME,
    'password': Config.LAB_DB_PASSWORD,
    'database': Config.LAB_DB_NAME,
    'charset': 'utf8mb4'
}

# ==================== SQL 查詢（大數據） ====================
query = """
SELECT 
    CASE 
        WHEN action.course_id IN ('770', '772', '775', '779') THEN '酷英AI英語聊天機器人'
        WHEN action.course_id IN ('941', '942', '943', '944') THEN '酷英篇章口說評測系統'
        WHEN action.course_id IN ('767', '768', '771', '773', '774', '776', '777', '778', '918', '919', '920', '921', '1358') THEN '酷英語音合成工具/語音合成工具'
        WHEN action.course_id IN ('841', '843', '922', '923', '924', '925') THEN '酷英AI寫作偵錯工具'
        WHEN action.course_id IN ('842', '844', '926', '927', '928', '930') THEN '多益寫作評估工具'
        WHEN action.course_id IN ('862', '864', '868', '937', '938', '939', '940', '1355', '1388') THEN 'Linggle Write'
        WHEN action.course_id IN ('989', '990', '991', '992') THEN '酷英教學＆學習工具區'
        WHEN action.course_id IN ('912', '915', '916', '917') THEN '酷英教師AI特助'
        WHEN action.course_id IN ('972', '973', '974', '975') THEN '酷英沉浸式閱讀工具'
    END AS ai_course_name,
    actor.category AS student_level,
    CASE 
        WHEN actor.filename REGEXP '^[0-9]{4}-[0-9]{2}-[0-9]{2}' THEN 
            SUBSTRING(actor.filename, 1, 7)
        WHEN actor.filename REGEXP '^[0-9]{6}' THEN 
            CONCAT(SUBSTRING(actor.filename, 1, 4), '-', SUBSTRING(actor.filename, 5, 2))
        ELSE SUBSTRING(actor.filename, 1, 7)
    END AS month,
    COUNT(DISTINCT actor.uid) AS user_count
FROM client_event
LEFT JOIN action ON client_event.action_id = action.id
LEFT JOIN actor ON client_event.actor_id = actor.id
WHERE 
    actor.filename IS NOT NULL
    AND action.course_id IN (
        '770', '772', '775', '779',
        '941', '942', '943', '944',
        '767', '768', '771', '773', '774', '776', '777', '778', '918', '919', '920', '921', '1358',
        '841', '843', '922', '923', '924', '925',
        '842', '844', '926', '927', '928', '930',
        '862', '864', '868', '937', '938', '939', '940', '1355', '1388',
        '989', '990', '991', '992',
        '912', '915', '916', '917',
        '972', '973', '974', '975'
    )
    AND actor.category IS NOT NULL
    AND actor.category IN ('國小', '國中', '普高', '技高', '大專')
GROUP BY ai_course_name, student_level, month
ORDER BY ai_course_name, month, 
    FIELD(actor.category, '國小', '國中', '普高', '技高', '大專')
"""

# ==================== 連接資料庫並查詢 ====================
print("正在連接 LAB 資料庫...")
print(f"連線資訊: {Config.LAB_DB_HOST}:{Config.LAB_DB_PORT}/{Config.LAB_DB_NAME}")

try:
    conn = mysql.connector.connect(**config)
    print("✅ 資料庫連線成功！")
    
    # 測試連線
    cursor = conn.cursor()
    cursor.execute("SELECT VERSION()")
    version = cursor.fetchone()
    print(f"MySQL 版本: {version[0]}")
    cursor.close()
    
    # 執行查詢
    print("正在執行查詢...")
    df = pd.read_sql(query, conn)
    print(f"✅ 查詢完成，共 {len(df)} 筆資料")
    
    # 儲存原始資料供後續使用（只保留月份的統計）
    stats_df = df[df['month'].notna() & (df['month'] != '')].copy()
    
except mysql.connector.Error as err:
    print(f"❌ 資料庫連線錯誤: {err}")
    sys.exit(1)
    
except Exception as e:
    print(f"❌ 其他錯誤: {e}")
    sys.exit(1)
    
finally:
    if 'conn' in locals() and conn.is_connected():
        conn.close()
        print("資料庫連線已關閉")

# ==================== 資料轉換 ====================
# 課程順序
course_order = [
    '酷英AI英語聊天機器人',
    '酷英篇章口說評測系統',
    '酷英語音合成工具/語音合成工具',
    '酷英AI寫作偵錯工具',
    '多益寫作評估工具',
    'Linggle Write',
    '酷英教學＆學習工具區',
    '酷英教師AI特助',
    '酷英沉浸式閱讀工具'
]

# 學制順序
level_order = ['國小', '國中', '普高', '技高', '大專']

# 建立完整的月份和學制組合
all_months = sorted(df['month'].unique())
all_combinations = [(month, level) for month in all_months for level in level_order]

# 建立透視表
pivot_df = df.pivot_table(
    index='ai_course_name',
    columns=['month', 'student_level'],
    values='user_count',
    fill_value=0,
    aggfunc='sum'
)

# 確保所有月份和學制組合都存在
for combo in all_combinations:
    if combo not in pivot_df.columns:
        pivot_df[combo] = 0

# 重新排序欄位
pivot_df = pivot_df[all_combinations]

# 重新排序課程
existing_courses = [c for c in course_order if c in pivot_df.index]
pivot_df = pivot_df.reindex(existing_courses)

# ==================== 匯出 Excel ====================
output_file = f'AI相關課程各學制的使用人次_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

wb = Workbook()
ws = wb.active
ws.title = '統計資料'

# 格式設定
header_font = Font(name='微軟正黑體', size=10, bold=True)
data_font = Font(name='微軟正黑體', size=10)
center_alignment = Alignment(horizontal='center', vertical='center')
left_alignment = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ==================== 寫入標題列 ====================
# 第1行：年月 (合併相同月份的欄位)
current_month = None
month_start_col = 2
col_idx = 2

for month, level in all_combinations:
    if month != current_month:
        # 新的月份開始
        if current_month is not None:
            # 合併前一個月份的欄位
            if col_idx - month_start_col > 1:
                ws.merge_cells(start_row=1, start_column=month_start_col, 
                              end_row=1, end_column=col_idx-1)
        current_month = month
        month_start_col = col_idx
        # 只在每個月份的第一個欄位寫入月份名稱
        cell = ws.cell(row=1, column=col_idx)
        cell.value = month
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    col_idx += 1

# 合併最後一個月份的欄位
if col_idx - month_start_col > 1:
    ws.merge_cells(start_row=1, start_column=month_start_col, 
                  end_row=1, end_column=col_idx-1)

# 第2行：學制
col_idx = 2
for month, level in all_combinations:
    cell = ws.cell(row=2, column=col_idx)
    cell.value = level
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border
    col_idx += 1

# 第1欄標題格式
for row in [1, 2]:
    cell = ws.cell(row=row, column=1)
    cell.font = header_font
    cell.alignment = center_alignment
    cell.border = thin_border

# ==================== 寫入課程資料 ====================
row_idx = 3
for course in existing_courses:
    # 課程名稱
    cell = ws.cell(row=row_idx, column=1)
    cell.value = course
    cell.font = data_font
    cell.alignment = left_alignment
    cell.border = thin_border
    
    # 數據
    col_idx = 2
    for month, level in all_combinations:
        cell = ws.cell(row=row_idx, column=col_idx)
        value = pivot_df.loc[course, (month, level)]
        cell.value = '' if value == 0 else int(value)
        cell.font = data_font
        cell.alignment = center_alignment
        cell.border = thin_border
        col_idx += 1
    
    row_idx += 1

# ==================== 調整欄寬 ====================
ws.column_dimensions['A'].width = 35
for col_idx in range(2, len(all_combinations) + 2):
    col_letter = get_column_letter(col_idx)
    ws.column_dimensions[col_letter].width = 8

# 凍結窗格
ws.freeze_panes = 'B3'

# 儲存檔案
wb.save(output_file)
print(f"✅ 匯出完成：{output_file}")
print(f"📊 課程數量：{len(existing_courses)}")
print(f"📅 月份範圍：{all_months[0]} ~ {all_months[-1]}")

# ==================== 寫入資料庫 ====================
if 'stats_df' in locals() and not stats_df.empty:
    # 準備資料（只保留月份的統計）
    batch_date = datetime.now().date()
    stats_data = stats_df.to_dict('records')
    
    # 寫入 LAB 資料庫
    print("\n正在連接 LAB 資料庫...")
    try:
        db_lab = ServiceLAB()
        print("✅ LAB 資料庫連線成功！")
        
        # 建立表（如果不存在）
        print("正在建立統計表...")
        db_lab.createTable_ai_course_stats()
        print("✅ 統計表建立/檢查完成")
        
        # 插入資料
        print(f"正在寫入統計資料到 LAB 資料庫（批次日期：{batch_date}）...")
        inserted_count_lab = db_lab.insert_ai_course_stats(stats_data, batch_date)
        print(f"✅ 成功寫入 {inserted_count_lab} 筆統計資料到 LAB 資料庫")
        
        db_lab.cursor.close()
        db_lab.db.close()
        
    except Exception as e:
        print(f"⚠️ 寫入 LAB 資料庫時發生錯誤: {e}")
        import traceback
        traceback.print_exc()
    
    # 寫入 NCU 資料庫
    print("\n正在連接 NCU 資料庫...")
    try:
        db_ncu = ServiceNCU()
        print("✅ NCU 資料庫連線成功！")
        
        # 建立表（如果不存在）
        print("正在建立統計表...")
        db_ncu.createTable_ai_course_stats()
        print("✅ 統計表建立/檢查完成")
        
        # 插入資料
        print(f"正在寫入統計資料到 NCU 資料庫（批次日期：{batch_date}）...")
        inserted_count_ncu = db_ncu.insert_ai_course_stats(stats_data, batch_date)
        print(f"✅ 成功寫入 {inserted_count_ncu} 筆統計資料到 NCU 資料庫")
        
        db_ncu.cursor.close()
        db_ncu.db.close()
        
    except Exception as e:
        print(f"⚠️ 寫入 NCU 資料庫時發生錯誤: {e}")
        import traceback
        traceback.print_exc()
else:
    print("⚠️ 沒有統計資料可寫入資料庫")
